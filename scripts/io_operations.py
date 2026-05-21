########################
###    Imports.py    ###
########################
# Contains functions for importing data in a specific format


##### Imports -----------
import os, sys
import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill
from Bio import SeqIO
from tqdm import tqdm
import numpy as np
import shutil
from manipulations import construct_presence_matrix, short_species_name, clean_dict_keys, clean_bact_names
from manipulations import binarize_host_range, hostrange_df_to_dict, hostrange_bact
import re
from paths import raw_data_path, data_prod_path

def call_hostrange_df(file : str, sheet_name : str = "sum_hostrange", TS : bool = False, sparse : bool = False, data2 : bool = False) -> list:
    """
    Used to retrieve the hostrange data of PFU between bacterias and bacteriophages, as well as bacteria strain lookup dictionary.

    Args:
        *fasta* (str): Full path to hostrange data sheet [excel]
        *sheet_name* (str): Sheet name in excel file
        *TS* (bool): Troubleshoot on or off (like verbose)
        *sparse* (bool): Whether to return a sparse hostrange DataFrame.
        *data2* (bool): Whether to use the second dataset format.
    Returns:
        *list[0]* (dict): bacteria lookup dictionary with bacteria species and strain names (strain names are unique)
        *list[1]* (pd.DataFrame): hostrange pd.Dataframe with bacteria strain as index and phage names as row columns

    """
    
    if data2:
        # Load the host range data from the Excel file
        host_range_df = pd.read_excel(
            file,
            sheet_name=sheet_name)

        host_range_df.rename(columns={"Unnamed: 0": "phage"}, inplace=True)

        # Creating bact species lookup dict
        bact_lookup = {}
        if "Species" not in host_range_df.columns: # Creating naive bact lookup due to missing species column
            for bact in host_range_df["phage"]:
                bact_lookup[bact] = bact
        else:
            for bact in host_range_df["phage"]:
                bact_lookup[bact] = host_range_df.loc[bact, "Species"]
        
    
    else:
        # Load the host range data from the Excel file
        host_range_df = pd.read_excel(
            file,
            sheet_name=sheet_name,
            header=1).drop(columns=["isolate ID", "Hostrange_analysis", "Phage"])

        # Create a lookup dictionary for bacteria species based on Seq ID - dict
        bact_lookup = host_range_df[["Seq ID", "Species"]].drop_duplicates(subset=['Seq ID']).set_index('Seq ID').to_dict()['Species']
        if TS: 
            print("Bacteria lookup dictionary created with", len(bact_lookup), "entries.")
            print(bact_lookup)

        # Make Seq ID to phage name mapping - pandas df
        host_range_df = host_range_df.drop(columns=["Species"]).set_index('Seq ID').rename_axis('phage').reset_index()
    
    # Shorten bacteria names from "J14_21_reoriented_merged.fasta" to "J14_21" for easier handling and matching with minhash sketch names
    # Use clean_bact_names function that takes a list of bacteria names as input
    host_range_df["phage"] = host_range_df["phage"].apply(lambda x: clean_bact_names(x))  # Keep only the first word of the phage name


    if TS: print(host_range_df.head())
    return [bact_lookup, host_range_df]

def load_minhash_sketches(in_dir : str, TS : bool = False, output_as_np : bool = False):
    """
    Load Sourmash Minhash sketches from a directory and concatenate them in a dictionary.

    Args:
        **in_dir** (str): Path to signatures
        **TS** (bool): Troubleshoot on or off
        **output_as_np** (bool): Output minhash vector (values in dict) as np.array rather than the default python list
    
    Returns:
        Dictionary of minhashes, with (extended) phage names / bacteria names as keys and it's minhash composition as a vector [list / np.array]
    
    """
    from sourmash import load_one_signature
    minhash_data = {}
    for filename in os.listdir(in_dir):
        if filename.endswith(('.sig', '.json')): # sourmash signature files
            filepath = os.path.join(in_dir, filename)
            if TS: print(f"filepath: {filepath}")
            try:
                # sourmash.load_signatures returns an iterator
                sig = load_one_signature(filepath)
                
                if not sig:
                    print(f"Warning: No signatures found in {filename}. Skipping.")
                    continue
                
                name = str(sig)
                if output_as_np:
                    hashes = np.array(sorted(sig.minhash.hashes.keys()))
                else:
                    hashes = sorted(sig.minhash.hashes.keys())
                minhash_data[name] = hashes

            except Exception as e:
                print(f"Error loading sketch file {filename}: {e}. Skipping.")

    return minhash_data

def presence_matrix(phage_minhash_dir : str = None, bact_minhash_dir : str = None, n = 0, k = 0, data2 : bool = False, reversecomp_data : bool = True, subset_features : list = None, TS : bool = False):
    """
    Create a binary presence matrix from minhash sketches.
    Combines the workflows of loading minhash sketches (load_minhash_sketches()) and generating a binary presence matrix (manipulations.construct_presence_matrix()).
    Returns the binary presence matrix along with the list of unique minhashes and entity-to-index mapping.
    The presence matrix contains both phage and bacteria rows, so check for host interactions for each pairwise mapping.
    If phage_minhash_dir or bact_minhash_dir are None, default paths will be used based on n and k values.

    Args:
        **phage_minhash_dir** (str): Directory containing phage minhash sketches for a specific run (n & kmer size)
        **bact_minhash_dir** (str): Directory containing bacteria minhash sketches for a specific run (n & kmer size)
        **n** (int | list): Number of minhashes used in the sketches
        **k** (int | list): Kmer size used in the sketches
        **data2** (bool): Whether to use the second dataset format (default: False, uses original dataset format)
        **reversecomp_data** (bool): Whether reverse complements were used in the minhash sketches
        **subset_features** (list): List of minhashes to subset the presence matrix to (default: None, uses all minhashes)
        **TS** (bool): Troubleshoot on or off
    
    Returns:
        **presence_matrix** (np.array): Binary presence matrix with shape (num_entities, num_unique_minhashes)\n
        **entity_to_index** (list): Mapping from entity names to their corresponding row indices in the presence matrix\n
        **minhash_to_index** (dict): Mapping from minhashes to their corresponding column indices in the presence matrix\n
        **phage_minhash_data** (dict): Dictionary of phage minhash sketches\n
        **bact_minhash_data** (dict): Dictionary of bacteria minhash sketches
    """
    # Handle differential n & k
    differential_nk = False
    try:
        bn, pn = n
        bk, pk = k
        if TS: print(bn, bk, pn, pk)
        differential_nk = True
    except:
        if type(n) != int or type(k) != int:
            raise ValueError("Please provide n and k arguments as either lists of size 2, or as integers")

    ### Load minhash sketches
    if phage_minhash_dir is None:
        if differential_nk:
            phage_minhash_dir = data_prod_path+f"SM_sketches/PhageMinhash_n{pn}_k{pk}/"
        else:
            phage_minhash_dir = data_prod_path+f"SM_sketches/PhageMinhash_n{n}_k{k}/"
    
    if bact_minhash_dir is None:
        if differential_nk:
            bact_minhash_dir = data_prod_path+f"SM_sketches/BactMinhash_n{bn}_k{bk}/"
        else:
            bact_minhash_dir = data_prod_path+f"SM_sketches/BactMinhash_n{n}_k{k}/"
    
    if reversecomp_data:
        phage_minhash_dir = phage_minhash_dir[:-1]+"_rev/"
        bact_minhash_dir = bact_minhash_dir[:-1]+"_rev/"
    
    if TS: 
        print(f"Loading phage minhash sketches from: {phage_minhash_dir}")
        print(f"Loading bacteria minhash sketches from: {bact_minhash_dir}")
    
    try:
        phage_minhash_data = load_minhash_sketches(phage_minhash_dir, 
                                                   TS=TS, output_as_np=True)
        bact_minhash_data = load_minhash_sketches(bact_minhash_dir, 
                                                  TS=TS, output_as_np=True)
    except Exception as e:
        print(f"Error loading minhash sketches: {e}")
        return None, None, None, None, None
    
    phage_minhash_data = clean_dict_keys(phage_minhash_data)
    bact_minhash_copy = bact_minhash_data.copy()
    for key in bact_minhash_copy.keys():
        new_key_list = clean_bact_names([key], data2=data2)
        bact_minhash_data[new_key_list[0]] = bact_minhash_data.pop(key)

    ### Extract unique minhashes
    unique_minhashes = set() #for both phage and bacteria combined

    for key, val in phage_minhash_data.items():
        for minhash in val:
            unique_minhashes.add(minhash)

    for key, val in bact_minhash_data.items():
        for minhash in val:
            unique_minhashes.add(minhash)

    if TS: print(f"\nUnique minhashes extracted with len: {len(unique_minhashes)}")

    ### Obtaining presence matrix
    # Combine both dictionaries into a single dictionary
    all_entities_minhashes = {**phage_minhash_data, **bact_minhash_data}

    # Get an ordered list of all entity names (will be the row labels)
    entity_names = sorted(list(all_entities_minhashes.keys()))

    # Get an ordered list of unique minhashes (will be the column labels)
    # Sorting is crucial to ensure consistency in the matrix columns
    if subset_features is not None:
        # Use only the provided subset_features, preserving their order and dropping missing ones
        # Ensure we only keep features that are in the extracted unique_minhashes
        filtered = [f for f in subset_features if f in unique_minhashes]
        missing = [f for f in subset_features if f not in unique_minhashes]
        if TS and missing:
            print(f"Warning: {len(missing)} subset_features not found and will be ignored.")
        sorted_minhashes = filtered
    else:
        sorted_minhashes = sorted(list(unique_minhashes))

    # Determine dimensions
    N = len(entity_names)  # Number of rows (entities)
    M = len(sorted_minhashes)  # Number of columns (unique minhashes)

    # Create a dictionary for quick lookup of minhash indices
    minhash_to_index = {minhash: i for i, minhash in enumerate(sorted_minhashes)}
    entity_to_index = {name: i for i, name in enumerate(entity_names)}

    # Initialize the binary matrix with all zeros
    binary_matrix = np.zeros((N, M), dtype=int)

    # Iterate through each entity (row)
    for i, entity_name in enumerate(entity_names):
        # Get the list of minhashes for the current entity
        minhashes_present = all_entities_minhashes[entity_name]

        # Iterate through the minhashes present in the entity
        for minhash in minhashes_present:
            # Skip minhashes not in the selected/available columns
            if minhash not in minhash_to_index:
                continue
            # Get the column index for this minhash
            j = minhash_to_index[minhash]

            # Set the corresponding cell in the matrix to 1
            binary_matrix[i, j] = 1

    ### Proof of concept outputs ###
    if TS:
        print("\nBinary presence matrix created with shape:", binary_matrix.shape)
        print("Sample rows (entities):", entity_names[:5])
        print("Sample columns (minhashes):", sorted_minhashes[:5])
    
    return binary_matrix, entity_to_index, minhash_to_index, phage_minhash_data, bact_minhash_data

def obtain_idx_to_entity_mapping(phage_minhash_data, bact_minhash_data, minhash_to_index, TS : bool = False):
    """
    Given the stored phage and bacteria minhash data, create a mapping from column index to entity name (phage or bacteria)

    """
    idx_to_entity = {}
    for name, val in phage_minhash_data.items():
        for minhash in val:
            idx_to_entity[minhash_to_index[minhash]] = name
    for name, val in bact_minhash_data.items():
        for minhash in val:
            idx_to_entity[minhash_to_index[minhash]] = name

    if TS: print(f"idx_to_entity mapping created with {len(idx_to_entity)} entries.")
    
    return idx_to_entity

def color_sheet_from_matrix(
        input_excel: str,
        sheet1_name: str,
        prediction_matrix_df: pd.DataFrame,
        output_excel: str = "colored_output.xlsx",
        excluded_bacteria : list = None,
        excluded_phages : list = None,
        TS: bool = False
    ):
    """
    Reads an Excel file, colors Sheet1 based on the corresponding values
    in torchMLP_prediction_matrix, and writes a new Excel file.
    Optionally highlights excluded bacteria/phage names in row/column labels.
    Assumptions:
    - Colors Sheet1 (F3:AB112) based on values in prediction matrix.
    - Sheet1 column names come from F2:AB2.
    - Sheet1 row names are in column B (rows 3+).
    """

    # ---- Load the Excel file ----
    df_sheet1 = pd.read_excel(input_excel, sheet_name=sheet1_name, header=None)

    # Extract row/column names
    row_names_sheet1 = df_sheet1.iloc[2:, 1]  # starting row 3
    col_start = 5      # F
    col_end = 27       # AB
    col_names_sheet1 = df_sheet1.iloc[1, col_start:col_end+1]
    if TS:
        print("Row names (Sheet1):", row_names_sheet1.tolist()[:5])
        print("Column names (Sheet1):", col_names_sheet1.tolist()[:5])

    #row_names_matrix = prediction_matrix_df.iloc[:, 0]       
    row_names_matrix = prediction_matrix_df.index     
    col_names_matrix = prediction_matrix_df.columns 
    if TS:
        print("Row names (Pred Matrix):", row_names_matrix.tolist()[:5])
        print("Column names (Pred Matrix):", col_names_matrix.tolist()[:5])

    # Convert matrix sheet into a lookup dict
    # dict[row_name][col_name] = 0 or 1
    matrix_dict = {
        row: prediction_matrix_df.loc[row].to_dict()
        for row in prediction_matrix_df.index
    }
    if TS:
        sample_row = row_names_matrix[0]
        print(f"\nSample row in matrix_dict: {sample_row} ->", matrix_dict[sample_row])

    # ---- Create new workbook for output ----
    wb = Workbook()
    ws = wb.active
    ws.title = sheet1_name

    # ---- Write Sheet1 values first ----
    for r in range(df_sheet1.shape[0]):
        for c in range(df_sheet1.shape[1]):
            ws.cell(r+1, c+1, df_sheet1.iat[r, c])

    # ---- Colors ----
    fill_1 = PatternFill(start_color="FFCCFFCC", end_color="FFCCFFCC", fill_type="solid")  # light green
    fill_0 = PatternFill(start_color="FFFFCCCC", end_color="FFFFCCCC", fill_type="solid")  # light red
    excluded_bact_fill = PatternFill(start_color="FFFFF2CC", end_color="FFFFF2CC", fill_type="solid")  # light yellow
    excluded_phage_fill = PatternFill(start_color="FFD9E1F2", end_color="FFD9E1F2", fill_type="solid")  # light blue

    excluded_bacteria_set = {str(x).strip() for x in (excluded_bacteria or [])}
    excluded_phages_set = {str(x).strip() for x in (excluded_phages or [])}

    # ---- Apply coloring in F3:AB112 or dynamically based on DataFrame ----
    for r_idx, row_name in enumerate(row_names_sheet1, start=3):

        # pandas index 5 == column F, but openpyxl column F == 6
        for c_index, col_name in enumerate(col_names_sheet1):

            excel_col = col_start + c_index + 1    # +1 for openpyxl 1-based indexing

            if col_name not in col_names_matrix:
                continue

            try:
                val = matrix_dict[row_name][col_name]
            except KeyError:
                continue

            cell = ws.cell(row=r_idx, column=excel_col)

            if val == 1:
                cell.fill = fill_1
            elif val == 0:
                cell.fill = fill_0

    # ---- Highlight excluded row/column labels (without overriding matrix-cell colors) ----
    for r_idx, row_name in enumerate(row_names_sheet1, start=3):
        if str(row_name).strip() in excluded_bacteria_set:
            ws.cell(row=r_idx, column=2).fill = excluded_bact_fill  # Column B (row label)

    for c_index, col_name in enumerate(col_names_sheet1):
        if str(col_name).strip() in excluded_phages_set:
            excel_col = col_start + c_index + 1
            ws.cell(row=2, column=excel_col).fill = excluded_phage_fill  # Header row

    if TS and (excluded_bacteria_set or excluded_phages_set):
        print(f"Excluded bacteria highlighted (row labels): {len(excluded_bacteria_set)}")
        print(f"Excluded phages highlighted (column headers): {len(excluded_phages_set)}")

    # ---- Save result ----
    wb.save(output_excel)
    print(f"\nSaved colored Excel as: {output_excel}")
